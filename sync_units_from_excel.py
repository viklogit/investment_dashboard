import sqlite3
import os
import openpyxl
from db import get_db, EXCEL_PATH

def sync():
    print(f"Syncing Units and Buy Prices from {EXCEL_PATH}...")
    if not os.path.exists(EXCEL_PATH):
        print("Excel file not found.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "Units" not in wb.sheetnames or "Buy Prices" not in wb.sheetnames:
        print("Required sheets 'Units' or 'Buy Prices' not found.")
        return

    conn = get_db()
    c = conn.cursor()

    # Get months map
    c.execute("SELECT id, label FROM months")
    month_map = {row['label']: row['id'] for row in c.fetchall()}
    
    # Get assets map
    c.execute("SELECT id, name FROM assets")
    asset_map = {row['name']: row['id'] for row in c.fetchall()}

    ws_units = wb["Units"]
    ws_prices = wb["Buy Prices"]

    # 1. Identify months from header (row 2)
    months = []
    for col in range(2, ws_units.max_column + 1):
        v = ws_units.cell(row=2, column=col).value
        if v and str(v).strip() and str(v).strip().upper() != "TOTAL":
            months.append((col, str(v).strip()))

    # 2. Process each asset row
    for row in range(3, ws_units.max_row + 1):
        asset_label = ws_units.cell(row=row, column=1).value
        if not asset_label or str(asset_label).strip().upper() in ("TOTAL", "PORTFOLIO TOTAL", "TOTAL INVESTED", "", "TOTAL PER MONTH"):
            continue
        
        asset_name = str(asset_label).strip()
        asset_id = asset_map.get(asset_name)
        
        if not asset_id:
            # Try to insert it if it doesn't exist? 
            # For now skip, assuming user creates assets first or initial import handled it.
            print(f"Asset '{asset_name}' not found in DB. Skipping.")
            continue

        for col, m_label in months:
            month_id = month_map.get(m_label)
            if not month_id:
                continue

            # Read Units
            units = ws_units.cell(row=row, column=col).value
            units = float(units) if units is not None else 0.0

            # Read Buy Price
            buy_price = ws_prices.cell(row=row, column=col).value
            
            # Fallback to market price if empty
            if (buy_price is None or buy_price == 0) and units > 0:
                c.execute("SELECT price FROM valuations WHERE asset_id=? AND month_id=?", (asset_id, month_id))
                val_row = c.fetchone()
                if val_row and val_row['price']:
                    buy_price = val_row['price']
                else:
                    buy_price = 0.0
            else:
                buy_price = float(buy_price) if buy_price is not None else 0.0

            amount = units * buy_price

            # Update contributions
            c.execute("SELECT id FROM contributions WHERE asset_id=? AND month_id=?", (asset_id, month_id))
            exist = c.fetchone()
            if exist:
                c.execute("""
                    UPDATE contributions 
                    SET units=?, buy_price=?, amount=? 
                    WHERE asset_id=? AND month_id=?
                """, (units, buy_price, amount, asset_id, month_id))
            else:
                c.execute("""
                    INSERT INTO contributions (asset_id, month_id, units, buy_price, amount)
                    VALUES (?, ?, ?, ?, ?)
                """, (asset_id, month_id, units, buy_price, amount))

    conn.commit()

    # 3. Recalculate valuations based on new units
    print("Recalculating units_held in valuations...")
    for aid in asset_map.values():
        running_units = 0.0
        # Order by month date_end to be sure
        c.execute("SELECT c.units, c.month_id, m.date_end FROM contributions c JOIN months m ON c.month_id = m.id WHERE c.asset_id=? ORDER BY m.date_end ASC", (aid,))
        rows = c.fetchall()
        for r in rows:
            running_units += r['units']
            mid = r['month_id']
            
            c.execute("SELECT price FROM valuations WHERE asset_id=? AND month_id=?", (aid, mid))
            v_row = c.fetchone()
            price = v_row['price'] if v_row and v_row['price'] else None
            
            if price:
                m_val = running_units * price
                c.execute("UPDATE valuations SET units_held=?, market_value=? WHERE asset_id=? AND month_id=?", (running_units, m_val, aid, mid))
            else:
                c.execute("UPDATE valuations SET units_held=? WHERE asset_id=? AND month_id=?", (running_units, aid, mid))

    conn.commit()
    conn.close()
    print("Sync complete.")

if __name__ == "__main__":
    sync()
