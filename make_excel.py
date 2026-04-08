import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INVESTMENTS = [
    "MSCI World (IWDA)", "S&P 500 (CSPX)", "Emerging Markets (EIMI)",
    "NASDAQ 100 (CNDX)", "European Stocks (MEUD)",
    "Apple (AAPL)", "NVIDIA (NVDA)", "Microsoft (MSFT)",
]
MONTHS = [
    "Jan 2024","Feb 2024","Mar 2024","Apr 2024","May 2024","Jun 2024",
    "Jul 2024","Aug 2024","Sep 2024","Oct 2024","Nov 2024","Dec 2024",
    "Jan 2025","Feb 2025","Mar 2025",
]
MONTHLY_DATA = {
    "MSCI World (IWDA)":       [300,300,300,300,300,300,300,300,300,300,300,300,350,350,350],
    "S&P 500 (CSPX)":          [200,200,200,200,200,200,200,200,200,200,200,200,200,200,200],
    "Emerging Markets (EIMI)": [100,100,100,0,0,100,100,0,100,100,0,100,100,0,100],
    "NASDAQ 100 (CNDX)":       [150,150,150,150,150,150,150,150,150,150,150,150,150,150,150],
    "European Stocks (MEUD)":  [100,100,100,100,100,100,100,100,100,100,100,100,0,0,100],
    "Apple (AAPL)":            [50,0,50,0,50,0,50,0,50,0,50,0,50,0,50],
    "NVIDIA (NVDA)":           [0,100,0,100,0,100,0,100,0,100,0,100,0,100,0],
    "Microsoft (MSFT)":        [75,75,75,75,75,75,75,75,75,75,75,75,75,75,75],
}

HEADER_FILL = PatternFill("solid", start_color="1B2A4A")
ALT_FILL    = PatternFill("solid", start_color="F0F4FA")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
TOTAL_FILL  = PatternFill("solid", start_color="E8F0FE")
GOLD_FILL   = PatternFill("solid", start_color="F4B942")
THIN        = Side(style="thin", color="C5D3E8")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT     = "#,##0.00 €"


def style_header(cell, text, gold=False):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = GOLD_FILL if gold else HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_label(cell, text, fill):
    cell.value = text
    cell.font = Font(bold=True, color="1B2A4A", size=10)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER


def style_data(cell, value, fill, num_fmt=NUM_FMT):
    cell.value = value if value else None
    cell.font = Font(color="2C3E50", size=10)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = num_fmt
    cell.border = BORDER


def style_total(cell, formula, gold=False):
    cell.value = formula
    cell.font = Font(bold=True, color="FFFFFF" if gold else "1B2A4A", size=10 if not gold else 11)
    cell.fill = GOLD_FILL if gold else TOTAL_FILL
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = NUM_FMT
    cell.border = BORDER


def setup_sheet(ws, title_text, n_months, has_totals=True):
    ws.freeze_panes = "B3"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20

    tc = n_months + 2 if has_totals else n_months + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=tc)
    t = ws.cell(row=1, column=1, value=title_text)
    t.font = Font(bold=True, color="FFFFFF", size=12)
    t.fill = HEADER_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    return tc


def build_monthly_sheet(ws, investments, months, data, title="Monthly Investment by Asset", show_totals=True, num_fmt=NUM_FMT):
    n = len(months)
    tc = setup_sheet(ws, title, n, has_totals=show_totals)

    style_header(ws.cell(row=2, column=1), "Asset")
    for j, m in enumerate(months, 2):
        style_header(ws.cell(row=2, column=j), m)
        ws.column_dimensions[get_column_letter(j)].width = 13
        
    if show_totals:
        style_header(ws.cell(row=2, column=tc), "TOTAL", gold=True)
        ws.column_dimensions[get_column_letter(tc)].width = 14

    for i, inv in enumerate(investments):
        row = i + 3
        fill = WHITE_FILL if i % 2 == 0 else ALT_FILL
        style_label(ws.cell(row=row, column=1), inv, fill)
        for j, val in enumerate(data[inv], 2):
            style_data(ws.cell(row=row, column=j), val, fill, num_fmt)
            
        if show_totals:
            last = get_column_letter(n + 1)
            style_total(ws.cell(row=row, column=tc), f"=SUM(B{row}:{last}{row})")

    if show_totals:
        tr = len(investments) + 3
        ws.cell(row=tr, column=1).value = "TOTAL PER MONTH"
        ws.cell(row=tr, column=1).font = Font(bold=True, color="1B2A4A", size=10)
        ws.cell(row=tr, column=1).fill = TOTAL_FILL
        ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=tr, column=1).border = BORDER
        for j in range(2, n + 2):
            col = get_column_letter(j)
            style_total(ws.cell(row=tr, column=j), f"=SUM({col}3:{col}{tr-1})")
        style_total(ws.cell(row=tr, column=tc),
                    f"=SUM({get_column_letter(tc)}3:{get_column_letter(tc)}{tr-1})", gold=True)


def build_cumulative_sheet(ws, investments, months):
    n = len(months)
    tc = setup_sheet(ws, "Cumulative Investment by Asset", n)

    style_header(ws.cell(row=2, column=1), "Asset")
    for j, m in enumerate(months, 2):
        style_header(ws.cell(row=2, column=j), m)
        ws.column_dimensions[get_column_letter(j)].width = 13
    style_header(ws.cell(row=2, column=tc), "LATEST", gold=True)
    ws.column_dimensions[get_column_letter(tc)].width = 14

    for i, inv in enumerate(investments):
        row = i + 3
        fill = WHITE_FILL if i % 2 == 0 else ALT_FILL
        style_label(ws.cell(row=row, column=1), inv, fill)
        for j in range(2, n + 2):
            col = get_column_letter(j)
            prev = get_column_letter(j - 1)
            if j == 2:
                formula = f"='Monthly Investment'!{col}{row}"
            else:
                formula = f"={prev}{row}+'Monthly Investment'!{col}{row}"
            c = ws.cell(row=row, column=j)
            c.value = formula
            c.font = Font(color="2C3E50", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="right")
            c.number_format = NUM_FMT
            c.border = BORDER
        last = get_column_letter(n + 1)
        style_total(ws.cell(row=row, column=tc), f"={last}{row}")

    tr = len(investments) + 3
    ws.cell(row=tr, column=1).value = "PORTFOLIO TOTAL"
    ws.cell(row=tr, column=1).font = Font(bold=True, color="1B2A4A", size=10)
    ws.cell(row=tr, column=1).fill = TOTAL_FILL
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=tr, column=1).border = BORDER
    for j in range(2, n + 2):
        col = get_column_letter(j)
        style_total(ws.cell(row=tr, column=j), f"=SUM({col}3:{col}{tr-1})")
    last = get_column_letter(n + 1)
    style_total(ws.cell(row=tr, column=tc), f"={last}{tr}", gold=True)


def build_change_sheet(ws, investments, months):
    n = len(months)
    tc = setup_sheet(ws, "Month-over-Month Investment Change", n)

    style_header(ws.cell(row=2, column=1), "Asset")
    for j, m in enumerate(months, 2):
        style_header(ws.cell(row=2, column=j), m)
        ws.column_dimensions[get_column_letter(j)].width = 13
    style_header(ws.cell(row=2, column=tc), "NET ADDED", gold=True)
    ws.column_dimensions[get_column_letter(tc)].width = 14

    for i, inv in enumerate(investments):
        row = i + 3
        fill = WHITE_FILL if i % 2 == 0 else ALT_FILL
        style_label(ws.cell(row=row, column=1), inv, fill)
        for j in range(2, n + 2):
            col = get_column_letter(j)
            prev = get_column_letter(j - 1)
            if j == 2:
                formula = f"='Monthly Investment'!{col}{row}"
            else:
                formula = f"='Monthly Investment'!{col}{row}-'Monthly Investment'!{prev}{row}"
            c = ws.cell(row=row, column=j)
            c.value = formula
            c.font = Font(color="2C3E50", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="right")
            c.number_format = NUM_FMT
            c.border = BORDER
        tc_col = get_column_letter(tc)
        monthly_tc = get_column_letter(n + 2)
        style_total(ws.cell(row=row, column=tc), f"='Monthly Investment'!{monthly_tc}{row}")

    tr = len(investments) + 3
    ws.cell(row=tr, column=1).value = "TOTAL INVESTED"
    ws.cell(row=tr, column=1).font = Font(bold=True, color="1B2A4A", size=10)
    ws.cell(row=tr, column=1).fill = TOTAL_FILL
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=tr, column=1).border = BORDER
    for j in range(2, n + 2):
        col = get_column_letter(j)
        style_total(ws.cell(row=tr, column=j), f"=SUM({col}3:{col}{tr-1})")
    style_total(ws.cell(row=tr, column=tc),
                f"=SUM({get_column_letter(tc)}3:{get_column_letter(tc)}{tr-1})", gold=True)


def create(path="investments.xlsx"):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Monthly Investment"
    build_monthly_sheet(ws1, INVESTMENTS, MONTHS, MONTHLY_DATA, "Monthly Investment by Asset")

    ws2 = wb.create_sheet("Cumulative Investment")
    build_cumulative_sheet(ws2, INVESTMENTS, MONTHS)

    ws3 = wb.create_sheet("Monthly Change")
    build_change_sheet(ws3, INVESTMENTS, MONTHS)

    zero_data = {inv: [0]*len(MONTHS) for inv in INVESTMENTS}
    ws4 = wb.create_sheet("Units")
    build_monthly_sheet(ws4, INVESTMENTS, MONTHS, zero_data, "Units Held by Asset", show_totals=True, num_fmt="#,##0.000")

    ws5 = wb.create_sheet("Buy Prices")
    build_monthly_sheet(ws5, INVESTMENTS, MONTHS, zero_data, "Average Buy Price per Unit", show_totals=False)

    wb.save(path)
    print(f"Saved: {path}")


if __name__ == "__main__":
    create()
