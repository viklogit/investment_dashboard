import sqlite3
import os
import openpyxl
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "portfolio.db")
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "investments.xlsx")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    
    # Create Assets Table
    c.execute("""
        CREATE TABLE IF NOT EXISTS assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            asset_type TEXT NOT NULL DEFAULT 'MANUAL',
            ticker TEXT,
            isin TEXT,
            currency TEXT DEFAULT 'EUR',
            buy_currency TEXT DEFAULT 'EUR',
            target_currency TEXT DEFAULT 'EUR',
            price_source TEXT DEFAULT 'manual'
        )
    """)
    
    # Create Months Table
    # label: "Jan 2024", date_end: "2024-01-31" (ISO string)
    c.execute("""
        CREATE TABLE IF NOT EXISTS months (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            label TEXT NOT NULL UNIQUE,
            date_end TEXT NOT NULL
        )
    """)
    
    # Create Contributions Table
    c.execute("""
        CREATE TABLE IF NOT EXISTS contributions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id INTEGER NOT NULL,
            month_id INTEGER NOT NULL,
            amount REAL NOT NULL DEFAULT 0.0,
            units REAL NOT NULL DEFAULT 0.0,
            buy_price REAL NOT NULL DEFAULT 0.0,
            FOREIGN KEY(asset_id) REFERENCES assets(id),
            FOREIGN KEY(month_id) REFERENCES months(id),
            UNIQUE(asset_id, month_id)
        )
    """)
    
    # Create Valuations Table V2
    c.execute("""
        CREATE TABLE IF NOT EXISTS valuations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id INTEGER NOT NULL,
            month_id INTEGER NOT NULL,
            market_value REAL NOT NULL DEFAULT 0.0,
            units_held REAL NOT NULL DEFAULT 0.0,
            price REAL,
            is_manual BOOLEAN NOT NULL DEFAULT 1,
            source TEXT DEFAULT 'manual',
            FOREIGN KEY(asset_id) REFERENCES assets(id),
            FOREIGN KEY(month_id) REFERENCES months(id),
            UNIQUE(asset_id, month_id)
        )
    """)
    
    conn.commit()
    conn.close()

def _parse_month_label(label):
    # Try to convert "Jan 2024" to "2024-01-31"
    try:
        dt = datetime.strptime(label.strip(), "%b %Y")
        # Go to next month, subtract 1 day
        next_month = dt.month % 12 + 1
        year = dt.year + (dt.month // 12)
        import calendar
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        return f"{dt.year}-{dt.month:02d}-{last_day:02d}"
    except Exception:
        return "2099-12-31" # Fallback

def import_from_excel_if_empty():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as count FROM assets")
    if c.fetchone()['count'] > 0:
        conn.close()
        return # Already populated
    
    if not os.path.exists(EXCEL_PATH):
        # We can create a default mock or just create the excel if it doesn't exist
        from make_excel import create
        create(EXCEL_PATH)
        
    print("Importing initial data from Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "Monthly Investment" not in wb.sheetnames:
        conn.close()
        return
        
    ws = wb["Monthly Investment"]
    months = []
    
    # Read months
    for col in range(2, ws.max_column):
        v = ws.cell(row=2, column=col).value
        if v and str(v).strip() and str(v).strip().upper() != "TOTAL":
            months.append(str(v).strip())
            
    # Insert months
    for m in months:
        date_end = _parse_month_label(m)
        try:
            c.execute("INSERT OR IGNORE INTO months (label, date_end) VALUES (?, ?)", (m, date_end))
        except sqlite3.IntegrityError:
            pass
            
    conn.commit()
    
    # Read assets and investments
    c.execute("SELECT id, label FROM months")
    month_map = {row['label']: row['id'] for row in c.fetchall()}
    
    for row in range(3, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        # Check end of table
        if not label or str(label).strip().upper() in ("TOTAL PER MONTH", "TOTAL", "", "PORTFOLIO TOTAL", "TOTAL INVESTED"):
            break
        
        asset_name = str(label).strip()
        
        # Determine basic type/ticker from name as a guess for v1
        asset_type = 'MANUAL'
        ticker = None
        price_source = 'manual'
        
        if "(" in asset_name and ")" in asset_name:
            import re
            match = re.search(r'\((.*?)\)', asset_name)
            if match:
                ticker_guess = match.group(1)
                # Just some rudimentary logic
                if ticker_guess in ('AAPL', 'NVDA', 'MSFT'):
                    asset_type = 'STOCK'
                    ticker = ticker_guess
                    price_source = 'auto'
                elif ticker_guess in ('IWDA', 'CSPX', 'EIMI', 'CNDX', 'MEUD'):
                    asset_type = 'ETF'
                    # appending .L or .AS depends on exchange, we'll just put the ticker
                    ticker = ticker_guess
                    price_source = 'auto'
        
        c.execute("""
            INSERT INTO assets (name, asset_type, ticker, price_source)
            VALUES (?, ?, ?, ?)
        """, (asset_name, asset_type, ticker, price_source))
        asset_id = c.lastrowid
        
        # Read contributions for this asset
        for j, m_label in enumerate(months, 2):
            val = ws.cell(row=row, column=j).value
            val = float(val) if val else 0.0
            month_id = month_map.get(m_label)
            if month_id:
                # Insert contribution
                c.execute("""
                    INSERT INTO contributions (asset_id, month_id, amount)
                    VALUES (?, ?, ?)
                """, (asset_id, month_id, val))
                
                # We need to initialize market_value in valuations.
                # Since we don't have historical prices, we'll initialize market_value to 
                # be roughly the cumulative sum so far just as a fake starting point? 
                # Or just exactly the cumulative cash so PnL starts at 0.
                pass
                
    conn.commit()
    
    # Initialize valuations with cumulative cash so PnL isn't negative out of nowhere
    c.execute("SELECT id FROM assets")
    all_assets = [r['id'] for r in c.fetchall()]
    c.execute("SELECT id FROM months ORDER BY id ASC")
    all_months = [r['id'] for r in c.fetchall()]
    
    for aid in all_assets:
        running = 0.0
        for mid in all_months:
            c.execute("SELECT amount FROM contributions WHERE asset_id=? AND month_id=?", (aid, mid))
            row = c.fetchone()
            val = row['amount'] if row else 0.0
            running += val
            
            # Insert valuation equal to running cash (yielding 0 PnL natively)
            c.execute("""
                INSERT INTO valuations (asset_id, month_id, market_value, is_manual, source)
                VALUES (?, ?, ?, 1, 'manual')
            """, (aid, mid, running))
            
    conn.commit()
    conn.close()
    print("Database built and seeded.")

if __name__ == "__main__":
    init_db()
    import_from_excel_if_empty()
